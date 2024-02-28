import pandas   as pd
from openpyxl import load_workbook
from  pykrx import stock
import time
import argparse
import mariadb
import ta

def get_args():
    """Get command-line arguments"""
    parser = argparse.ArgumentParser(
        description="Major investor Buying tickers",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument('basedate', metavar='basedate', help='basedate')
    parser.add_argument('-r', action='store_true',help='correlation coefficient')
    parser.add_argument('-p', action='store_true',help='buy power 순매수/시총')

    return parser.parse_args()

def get_connection() :
    """
    마리아DB 접속 Connection 을 획득한다.
    """
    conn = None
    try:
        conn = mariadb.connect(
            user="root",
            password="123",
            host="localhost",
            port=3306,
            database="home"
        )
    except mariadb.Error as e:
        print(f"Error connecting to MariaDB Platform: {e}")
    finally :
        return conn
    
def  get_tickers(base_date) :
    """
        get_market_ticker_list 로 종목리스트를 작성하여 리턴한다.
    """
    tickers=[]
    tickers_list = stock.get_market_ticker_list(base_date,market='KOSPI')
    for ticker in tickers_list :
        tickers.append((ticker,stock.get_market_ticker_name(ticker),'KOSPI'))

    tickers_list = stock.get_market_ticker_list(base_date,market='KOSDAQ')
    for ticker in tickers_list :
        tickers.append((ticker,stock.get_market_ticker_name(ticker),'KOSDAQ'))
    
    return tickers

def corr(base_date) :
    sql = """
        SELECT a.code as 코드
            ,b.name as 종목명
            ,a.date as 날짜
            ,a.close as  종가
            ,c.ts + c.sm + c.yk  AS 기관
            ,c.fr AS 외국인
            
        FROM   daily_price   a

        JOIN   company_info b
                ON  A.CODE = B.CODE

        JOIN   daily_agent  c 
                ON  a.code = c.code
                AND a.date = c.date
                
        WHERE  a.date >= date_add(now(), INTERVAL -4 MONTH)  
    """
    conn = get_connection()
    df  = pd.read_sql(sql, conn)
    df['기관누적'] = df['기관'].cumsum()
    df['외국인누적'] = df['외국인'].cumsum()
    df_corr = df[["코드","종목명","날짜","종가","기관누적","외국인누적"]]
    df_corr.dropna(inplace=True)
    df_corr['MA5_기관누적']=ta.trend.SMAIndicator(close=df_corr["기관누적"],window=5).sma_indicator()
    df_corr['MA5_외인누적']=ta.trend.SMAIndicator(close=df_corr["외국인누적"],window=5).sma_indicator()

    df_corr_기관 = df_corr.groupby(["코드","종목명"]).apply(lambda p : p['종가'].corr(p['기관누적'])).reset_index()
    df_corr_외인 = df_corr.groupby(["코드","종목명"]).apply(lambda p : p['종가'].corr(p['외국인누적'])).reset_index()
    df_corr_기관.rename(columns={0:"Corr기관"},inplace=True)
    df_corr_외인.rename(columns={0:"Corr외인"},inplace=True)

    df_cap = stock.get_market_cap(base_date).reset_index()
    df_cap.rename(columns={'티커':"코드"}, inplace=True)
    df_cap = df_cap[["코드","시가총액"]]

    df_corr = df_corr[df_corr["날짜"]==base_date]
    df_corr = df_corr.merge(df_corr_기관, on=["코드","종목명"],how="left")
    df_corr = df_corr.merge(df_corr_외인, on=["코드","종목명"],how="left")
    func = lambda r : "기관" if r["Corr기관"] > r["Corr외인"] and r["Corr기관"] > 0.8  \
        else ("외인" if  r["Corr기관"] < r["Corr외인"] and r["Corr외인"] > 0.8 else "")
    df_corr["주도주체"] = df_corr.apply(func , axis=1) 
    df_corr = df_corr.merge(df_cap, on=["코드"],how="left")
    df_corr.dropna(inplace=True)
    print(f"correlation coefficient dataframe created!!! {len(df_corr)} 건")
    result = df_corr.values.tolist()

    wb = load_workbook("Major.xlsx")
    del  wb["R"]
    wb.create_sheet("R")
    ws = wb["R"]
    ws.append(["코드","종목명","날짜","종가","기관누적","외국인누적","기관누적MA5","외인누적MA5","상관계수기관","상관계수외인","주도주체","시총"])
    for r in result :
        ws.append(r)

    wb.save("Major.xlsx")    
    wb.close()


def simulation(base_date) :
    result = []
    df = pd.read_excel("Major.xlsx", sheet_name="R")
    r_row_count = len(df)
    df.dropna(inplace=True)             # 주도주체가 "기관", "외국인" 중 하나가 들어 있지 않으면 데이터가 제거됨
    tickers = df[df["시총"] > 100_000_000_000][["코드","종목명","주도주체","시총"]].values.tolist()
    df =None
    row_count = 1
    for ticker in tickers :
        try :
            time.sleep(1)
            df_investor = stock.get_market_trading_value_by_investor(base_date, base_date, ticker[0])
            major_cap = int(ticker[3])
            major_investor = df_investor[df_investor.index.isin(["투신","사모","연기금 등","외국인"])]["순매수"].sum()
            purchase_size = major_investor/major_cap * 100
        except ValueError :
            print(f"[df fetch :{base_date=} {ticker[0]=} ] ValueError Continue...")
            continue

        if df_investor.empty:
            continue

        if  purchase_size > 0.2 :   # 시총천억이상 매수강도 0.2이상
            print(f'{ticker[1]}({ticker[0]}) : {purchase_size:.2f}')
            result.append([ticker[0],ticker[1],purchase_size
                           ,f"=VLOOKUP($A{row_count+1},'R'!$A$2:$L${r_row_count+1},9,FALSE)"
                           ,f"=VLOOKUP($A{row_count+1},'R'!$A$2:$L${r_row_count+1},10,FALSE)"
                           ,f"=VLOOKUP($A{row_count+1},'R'!$A$2:$L${r_row_count+1},11,FALSE)"
                           ])
            row_count += 1

    wb = load_workbook("Major.xlsx")
    ws = wb.create_sheet(base_date)

    ws.append(["종목코드","종목명","매수강도","R기관","R외인","주도주체"])
    for r in result :
        ws.append(r)

    wb.save("Major.xlsx")          
    wb.close()

    return result

def main() :
    args = get_args()
    base_date = args.basedate
    if args.r :
        corr(base_date)
    elif args.p :
        simulation(base_date)

if __name__=="__main__" :
    main()